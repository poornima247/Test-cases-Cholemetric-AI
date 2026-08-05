import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def load_results():
    try:
        with open("security-testing/results.json", "r") as f:
            return json.load(f)
    except FileNotFoundError:
        print("results.json not found. Run run_security_tests.py first.")
        return None

def generate_excel_reports():
    data = load_results()
    if not data:
        return

    os.makedirs("security-testing/reports", exist_ok=True)
    
    wb = Workbook()
    
    # Sheet 1: Security Findings (Sample Data for structure)
    ws1 = wb.active
    ws1.title = "Security Findings"
    headers1 = ["Finding ID", "Severity", "Vuln Type", "CWE", "OWASP", "File", "Endpoint", "Description"]
    ws1.append(headers1)
    
    sample_findings = [
        ["SEC-001", "CRITICAL", "Hardcoded Credentials", "CWE-798", "A07:2021", "db.php", "All", "Root user with empty password"],
        ["SEC-002", "CRITICAL", "Broken Authentication", "CWE-306", "A07:2021", "Multiple", "Multiple", "No token/session validation"],
        ["SEC-003", "HIGH", "CORS Misconfig", "CWE-942", "A01:2021", "Multiple", "Multiple", "Wildcard Access-Control-Allow-Origin"],
        ["SEC-004", "HIGH", "Arbitrary File Upload", "CWE-434", "A04:2021", "analyze.php", "/analyze.php", "Missing MIME validation"]
    ]
    for row in sample_findings:
        ws1.append(row)

    # Apply some basic colors
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    high_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value == "CRITICAL":
                cell.fill = critical_fill
            elif cell.value == "HIGH":
                cell.fill = high_fill

    # Sheet 2: Test Cases
    ws2 = wb.create_sheet(title="Test Cases")
    headers2 = ["ID", "Category", "Title", "Severity", "Status", "Actual Result"]
    ws2.append(headers2)
    
    for tc in data.get("results", []):
        ws2.append([
            tc.get("id"),
            tc.get("category"),
            tc.get("title"),
            tc.get("severity"),
            tc.get("status"),
            tc.get("actual_result", "")
        ])

    # Save master workbook
    wb.save("security-testing/reports/findings.xlsx")
    print("Excel reports generated successfully in security-testing/reports/findings.xlsx")

if __name__ == "__main__":
    generate_excel_reports()
