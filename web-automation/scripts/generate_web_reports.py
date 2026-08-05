import os
import sys
import glob
import json
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_surefire_reports(base_path):
    """Parse TestNG/Maven surefire XML reports and return detailed test results."""
    xml_files = glob.glob(os.path.join(base_path, "target", "surefire-reports", "TEST-*.xml"))
    
    total, passed, failed, skipped = 0, 0, 0, 0
    test_cases = []
    suite_stats = {}
    
    for xf in xml_files:
        try:
            tree = ET.parse(xf)
            root = tree.getroot()
            suite_name = root.attrib.get('name', os.path.basename(xf))
            
            s_total = int(root.attrib.get('tests', 0))
            s_failed = int(root.attrib.get('failures', 0)) + int(root.attrib.get('errors', 0))
            s_skipped = int(root.attrib.get('skipped', 0))
            s_passed = s_total - s_failed - s_skipped
            s_time = float(root.attrib.get('time', 0))
            
            total += s_total
            failed += s_failed
            skipped += s_skipped
            
            suite_stats[suite_name] = {
                'total': s_total, 'passed': s_passed,
                'failed': s_failed, 'skipped': s_skipped,
                'time': round(s_time, 2)
            }
            
            for tc in root.findall('testcase'):
                tc_name = tc.attrib.get('name', 'Unknown')
                tc_class = tc.attrib.get('classname', '').split('.')[-1]
                tc_time = float(tc.attrib.get('time', 0))
                
                failure_el = tc.find('failure')
                error_el = tc.find('error')
                skip_el = tc.find('skipped')
                
                if failure_el is not None or error_el is not None:
                    status = 'FAILED'
                    msg_el = failure_el if failure_el is not None else error_el
                    message = msg_el.attrib.get('message', '')[:200]
                elif skip_el is not None:
                    status = 'SKIPPED'
                    message = skip_el.attrib.get('message', 'Skipped')[:200]
                else:
                    status = 'PASSED'
                    message = ''
                
                test_cases.append({
                    'id': f"TC_{len(test_cases)+1:03d}",
                    'name': tc_name,
                    'class': tc_class,
                    'suite': suite_name,
                    'status': status,
                    'duration_ms': round(tc_time * 1000, 2),
                    'message': message
                })
        except Exception as e:
            print(f"  Warning: Could not parse {xf}: {e}")
    
    passed = total - failed - skipped
    return total, passed, failed, skipped, test_cases, suite_stats


def generate_html_report(out_dir, total, passed, failed, skipped, test_cases, suite_stats, run_time):
    """Generate a rich HTML execution report with charts and tables."""
    pass_rate = f"{(passed/total*100):.1f}" if total > 0 else "0.0"
    status_color = "#00e676" if failed == 0 else ("#ff9800" if failed/total < 0.1 else "#f44336")
    
    # Build test rows HTML
    rows_html = ""
    for tc in test_cases:
        color = {"PASSED": "#00e676", "FAILED": "#f44336", "SKIPPED": "#ff9800"}.get(tc['status'], "#888")
        badge = {"PASSED": "✅", "FAILED": "❌", "SKIPPED": "⏭️"}.get(tc['status'], "?")
        msg = f"<br><small style='color:#aaa'>{tc['message']}</small>" if tc['message'] else ""
        rows_html += f"""
        <tr>
            <td style='color:#90caf9'>{tc['id']}</td>
            <td>{tc['name']}{msg}</td>
            <td style='color:#b39ddb'>{tc['class']}</td>
            <td><span style='color:{color}'>{badge} {tc['status']}</span></td>
            <td style='color:#80cbc4'>{tc['duration_ms']} ms</td>
        </tr>"""

    # Build suite rows HTML
    suite_rows = ""
    for name, stats in suite_stats.items():
        sr = f"{(stats['passed']/stats['total']*100):.0f}%" if stats['total'] > 0 else "N/A"
        suite_rows += f"""
        <tr>
            <td style='color:#90caf9'>{name.split('.')[-1]}</td>
            <td>{stats['total']}</td>
            <td style='color:#00e676'>{stats['passed']}</td>
            <td style='color:#f44336'>{stats['failed']}</td>
            <td style='color:#ff9800'>{stats['skipped']}</td>
            <td>{stats['time']}s</td>
            <td>{sr}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cholemetric Web E2E Test Report</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Inter', sans-serif; background: #0a0e1a; color: #e0e0e0; min-height: 100vh; }}
        .header {{ background: linear-gradient(135deg, #1565c0 0%, #0d47a1 50%, #006064 100%);
                   padding: 40px; text-align: center; box-shadow: 0 4px 20px rgba(0,0,0,0.5); }}
        .header h1 {{ font-size: 2.5rem; font-weight: 700; color: #fff;
                      text-shadow: 0 2px 10px rgba(0,0,0,0.3); }}
        .header p {{ color: rgba(255,255,255,0.8); margin-top: 8px; font-size: 1rem; }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 30px; }}
        .metrics-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                         gap: 20px; margin: 30px 0; }}
        .metric-card {{ background: linear-gradient(145deg, #1a1f35, #141828);
                        border-radius: 16px; padding: 24px; text-align: center;
                        border: 1px solid rgba(255,255,255,0.08);
                        box-shadow: 0 8px 32px rgba(0,0,0,0.3);
                        transition: transform 0.2s; }}
        .metric-card:hover {{ transform: translateY(-4px); }}
        .metric-value {{ font-size: 3rem; font-weight: 700; line-height: 1; }}
        .metric-label {{ font-size: 0.85rem; color: #90a4ae; margin-top: 8px; text-transform: uppercase; letter-spacing: 1px; }}
        .section {{ background: #141828; border-radius: 16px; padding: 28px; margin: 24px 0;
                    border: 1px solid rgba(255,255,255,0.06); }}
        .section h2 {{ font-size: 1.3rem; font-weight: 600; color: #90caf9; margin-bottom: 20px; }}
        .charts-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin: 24px 0; }}
        .chart-box {{ background: #141828; border-radius: 16px; padding: 24px;
                      border: 1px solid rgba(255,255,255,0.06); }}
        .chart-box h3 {{ color: #90caf9; margin-bottom: 16px; font-size: 1rem; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th {{ background: #1a2744; color: #90caf9; padding: 12px 16px; text-align: left;
              font-size: 0.85rem; text-transform: uppercase; letter-spacing: 0.5px; }}
        td {{ padding: 10px 16px; border-bottom: 1px solid rgba(255,255,255,0.05); font-size: 0.9rem; }}
        tr:hover td {{ background: rgba(255,255,255,0.03); }}
        .badge {{ padding: 4px 12px; border-radius: 20px; font-size: 0.75rem; font-weight: 600; }}
        .pass-rate {{ font-size: 5rem; font-weight: 800; color: {status_color}; }}
        .info-row {{ display: flex; flex-wrap: wrap; gap: 20px; margin-bottom: 20px; }}
        .info-item {{ background: rgba(255,255,255,0.05); padding: 10px 16px; border-radius: 8px; font-size: 0.9rem; }}
        .info-item span {{ color: #90caf9; font-weight: 600; }}
        footer {{ text-align: center; padding: 30px; color: #546e7a; font-size: 0.85rem; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🧪 Cholemetric Web E2E Test Report</h1>
        <p>Generated: {run_time} | Framework: Selenium 4 + TestNG 7 + Maven</p>
    </div>
    
    <div class="container">
        <div class="info-row">
            <div class="info-item">🔢 Total Test Cases: <span>{total}</span></div>
            <div class="info-item">📅 Run Date: <span>{run_time}</span></div>
            <div class="info-item">🌐 Target: <span>Cholemetric GitHub Pages</span></div>
            <div class="info-item">⚙️ Browser: <span>Chrome (Headless)</span></div>
        </div>

        <div class="metrics-grid">
            <div class="metric-card">
                <div class="metric-value" style="color:#64b5f6">{total}</div>
                <div class="metric-label">📋 Total</div>
            </div>
            <div class="metric-card">
                <div class="metric-value" style="color:#00e676">{passed}</div>
                <div class="metric-label">✅ Passed</div>
            </div>
            <div class="metric-card">
                <div class="metric-value" style="color:#f44336">{failed}</div>
                <div class="metric-label">❌ Failed</div>
            </div>
            <div class="metric-card">
                <div class="metric-value" style="color:#ff9800">{skipped}</div>
                <div class="metric-label">⏭️ Skipped</div>
            </div>
            <div class="metric-card">
                <div class="pass-rate">{pass_rate}%</div>
                <div class="metric-label">🎯 Pass Rate</div>
            </div>
        </div>

        <div class="charts-row">
            <div class="chart-box">
                <h3>📊 Test Result Distribution</h3>
                <canvas id="pieChart" height="280"></canvas>
            </div>
            <div class="chart-box">
                <h3>📈 Suite-wise Breakdown</h3>
                <canvas id="barChart" height="280"></canvas>
            </div>
        </div>

        <div class="section">
            <h2>📦 Suite Summary</h2>
            <table>
                <thead><tr>
                    <th>Suite</th><th>Total</th><th>Passed</th>
                    <th>Failed</th><th>Skipped</th><th>Time</th><th>Pass Rate</th>
                </tr></thead>
                <tbody>{suite_rows if suite_rows else '<tr><td colspan="7" style="text-align:center;color:#90a4ae">No suite data available</td></tr>'}</tbody>
            </table>
        </div>

        <div class="section">
            <h2>📋 Detailed Test Cases</h2>
            <table>
                <thead><tr>
                    <th>ID</th><th>Test Name</th><th>Class</th>
                    <th>Status</th><th>Duration</th>
                </tr></thead>
                <tbody>{rows_html if rows_html else '<tr><td colspan="5" style="text-align:center;color:#90a4ae">No individual test data available — check surefire XML reports</td></tr>'}</tbody>
            </table>
        </div>
    </div>

    <footer>🔬 Cholemetric AI — Automated E2E Test Suite | Generated by Cholemetric CI/CD Pipeline</footer>

    <script>
    const pieCtx = document.getElementById('pieChart');
    new Chart(pieCtx, {{
        type: 'doughnut',
        data: {{
            labels: ['Passed', 'Failed', 'Skipped'],
            datasets: [{{
                data: [{passed}, {failed}, {skipped}],
                backgroundColor: ['#00e676', '#f44336', '#ff9800'],
                borderWidth: 0,
                hoverOffset: 8
            }}]
        }},
        options: {{
            responsive: true,
            plugins: {{
                legend: {{ position: 'bottom', labels: {{ color: '#e0e0e0', padding: 20 }} }}
            }}
        }}
    }});

    const suiteNames = {json.dumps(list(suite_stats.keys()))};
    const shortNames = suiteNames.map(n => n.split('.').pop().substring(0, 20));
    const passedData = {json.dumps([s['passed'] for s in suite_stats.values()])};
    const failedData = {json.dumps([s['failed'] for s in suite_stats.values()])};

    const barCtx = document.getElementById('barChart');
    new Chart(barCtx, {{
        type: 'bar',
        data: {{
            labels: shortNames,
            datasets: [
                {{ label: 'Passed', data: passedData, backgroundColor: '#00e676' }},
                {{ label: 'Failed', data: failedData, backgroundColor: '#f44336' }}
            ]
        }},
        options: {{
            responsive: true,
            scales: {{
                x: {{ stacked: true, ticks: {{ color: '#90a4ae' }}, grid: {{ color: 'rgba(255,255,255,0.05)' }} }},
                y: {{ stacked: true, ticks: {{ color: '#90a4ae' }}, grid: {{ color: 'rgba(255,255,255,0.05)' }} }}
            }},
            plugins: {{
                legend: {{ labels: {{ color: '#e0e0e0' }} }}
            }}
        }}
    }});
    </script>
</body>
</html>"""
    
    with open(os.path.join(out_dir, "HTML", "execution-report.html"), "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✅ HTML report generated: {len(test_cases)} test cases")


def generate_excel_report(out_dir, total, passed, failed, skipped, test_cases, suite_stats, run_time):
    """Generate comprehensive Excel reports."""
    
    # === Main Automation Report ===
    wb = Workbook()
    
    # Remove default sheet
    default = wb.active
    
    # --- Summary Sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header style
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    pass_fill = PatternFill(start_color="00C853", end_color="00C853", fill_type="solid")
    fail_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    skip_fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
    
    summary_data = [
        ("Metric", "Value"),
        ("Report Generated", run_time),
        ("Total Tests", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Skipped", skipped),
        ("Pass Rate", f"{(passed/total*100):.1f}%" if total > 0 else "N/A"),
        ("Framework", "Selenium 4 + TestNG 7"),
        ("Browser", "Chrome (Headless)"),
        ("Target", "Cholemetric GitHub Pages"),
    ]
    
    for r, (key, val) in enumerate(summary_data, 1):
        ws_summary.cell(row=r, column=1, value=key).font = Font(bold=True, color="90CAF9")
        ws_summary.cell(row=r, column=2, value=str(val))
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    
    # --- All Tests Sheet ---
    ws_all = wb.create_sheet("All Tests")
    headers = ["Test ID", "Test Name", "Class", "Suite", "Status", "Duration (ms)", "Error Message"]
    for col, header in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws_all.column_dimensions[get_column_letter(col)].width = [12, 50, 30, 40, 12, 15, 60][col-1]
    
    for r, tc in enumerate(test_cases, 2):
        ws_all.cell(row=r, column=1, value=tc['id'])
        ws_all.cell(row=r, column=2, value=tc['name'])
        ws_all.cell(row=r, column=3, value=tc['class'])
        ws_all.cell(row=r, column=4, value=tc['suite'])
        status_cell = ws_all.cell(row=r, column=5, value=tc['status'])
        if tc['status'] == 'PASSED':
            status_cell.fill = pass_fill
            status_cell.font = Font(color="FFFFFF", bold=True)
        elif tc['status'] == 'FAILED':
            status_cell.fill = fail_fill
            status_cell.font = Font(color="FFFFFF", bold=True)
        else:
            status_cell.fill = skip_fill
        ws_all.cell(row=r, column=6, value=tc['duration_ms'])
        ws_all.cell(row=r, column=7, value=tc['message'])
    
    # --- Passed Sheet ---
    ws_pass = wb.create_sheet("Passed")
    for col, h in enumerate(headers, 1):
        cell = ws_pass.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        cell.font = header_font
    for r, tc in enumerate([t for t in test_cases if t['status'] == 'PASSED'], 2):
        ws_pass.cell(row=r, column=1, value=tc['id'])
        ws_pass.cell(row=r, column=2, value=tc['name'])
        ws_pass.cell(row=r, column=3, value=tc['class'])
        ws_pass.cell(row=r, column=4, value=tc['suite'])
        ws_pass.cell(row=r, column=5, value=tc['status'])
        ws_pass.cell(row=r, column=6, value=tc['duration_ms'])
    
    # --- Failed Sheet ---
    ws_fail = wb.create_sheet("Failed")
    for col, h in enumerate(headers, 1):
        cell = ws_fail.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
        cell.font = header_font
    for r, tc in enumerate([t for t in test_cases if t['status'] == 'FAILED'], 2):
        ws_fail.cell(row=r, column=1, value=tc['id'])
        ws_fail.cell(row=r, column=2, value=tc['name'])
        ws_fail.cell(row=r, column=3, value=tc['class'])
        ws_fail.cell(row=r, column=4, value=tc['suite'])
        ws_fail.cell(row=r, column=5, value=tc['status'])
        ws_fail.cell(row=r, column=6, value=tc['duration_ms'])
        ws_fail.cell(row=r, column=7, value=tc['message'])
    
    # --- Suite Stats Sheet ---
    ws_suite = wb.create_sheet("Suite Stats")
    suite_headers = ["Suite", "Total", "Passed", "Failed", "Skipped", "Time (s)", "Pass Rate"]
    for col, h in enumerate(suite_headers, 1):
        cell = ws_suite.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for r, (name, stats) in enumerate(suite_stats.items(), 2):
        ws_suite.cell(row=r, column=1, value=name.split('.')[-1])
        ws_suite.cell(row=r, column=2, value=stats['total'])
        ws_suite.cell(row=r, column=3, value=stats['passed'])
        ws_suite.cell(row=r, column=4, value=stats['failed'])
        ws_suite.cell(row=r, column=5, value=stats['skipped'])
        ws_suite.cell(row=r, column=6, value=stats['time'])
        pr = f"{(stats['passed']/stats['total']*100):.1f}%" if stats['total'] > 0 else "N/A"
        ws_suite.cell(row=r, column=7, value=pr)
    
    excel_path = os.path.join(out_dir, "Excel", "Automation_Test_Report.xlsx")
    wb.save(excel_path)
    print(f"  ✅ Excel report saved: {excel_path}")
    
    # === Separate Failed report ===
    wb_fail = Workbook()
    ws_f = wb_fail.active
    ws_f.title = "Failed Tests"
    for col, h in enumerate(headers, 1):
        cell = ws_f.cell(row=1, column=col, value=h)
        cell.fill = fail_fill
        cell.font = header_font
    failed_cases = [t for t in test_cases if t['status'] == 'FAILED']
    for r, tc in enumerate(failed_cases, 2):
        ws_f.cell(row=r, column=1, value=tc['id'])
        ws_f.cell(row=r, column=2, value=tc['name'])
        ws_f.cell(row=r, column=3, value=tc['class'])
        ws_f.cell(row=r, column=4, value=tc['suite'])
        ws_f.cell(row=r, column=5, value=tc['status'])
        ws_f.cell(row=r, column=6, value=tc['duration_ms'])
        ws_f.cell(row=r, column=7, value=tc['message'])
    wb_fail.save(os.path.join(out_dir, "Excel", "Failed_Test_Cases.xlsx"))


def main():
    print("🚀 Generating Cholemetric Web E2E Reports...")
    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_dir = os.path.join(base_path, "Test Results")
    
    for sub in ["HTML", "Excel", "JSON", "Summary"]:
        os.makedirs(os.path.join(out_dir, sub), exist_ok=True)
    
    run_time = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    
    print("📊 Parsing surefire XML reports...")
    total, passed, failed, skipped, test_cases, suite_stats = parse_surefire_reports(base_path)
    print(f"  Found: {total} tests — ✅{passed} passed, ❌{failed} failed, ⏭️{skipped} skipped")
    
    print("📝 Generating HTML report...")
    generate_html_report(out_dir, total, passed, failed, skipped, test_cases, suite_stats, run_time)
    
    print("📊 Generating Excel reports...")
    generate_excel_report(out_dir, total, passed, failed, skipped, test_cases, suite_stats, run_time)
    
    print("📦 Writing JSON execution results...")
    json_results = {
        "total": total, "passed": passed, "failed": failed, "skipped": skipped,
        "pass_rate": round(passed/total*100, 2) if total > 0 else 0,
        "run_time": run_time,
        "suite_stats": suite_stats,
        "test_cases": test_cases[:100]  # Limit to first 100 for JSON size
    }
    with open(os.path.join(out_dir, "JSON", "execution-results.json"), "w") as f:
        json.dump(json_results, f, indent=2)
    
    print("📋 Writing Summary Markdown...")
    pr = f"{(passed/total*100):.1f}%" if total > 0 else "0.0%"
    md = f"""## 🧪 Cholemetric Web E2E Test Summary

| Metric | Value |
|--------|-------|
| **Total Tests** | {total} |
| **✅ Passed** | {passed} |
| **❌ Failed** | {failed} |
| **⏭️ Skipped** | {skipped} |
| **Pass Rate** | {pr} |
| **Run Time** | {run_time} |
"""
    
    summary_path = os.path.join(out_dir, "Summary", "summary.md")
    with open(summary_path, "w") as f:
        f.write(md)
    
    # Also write to root Test Results for workflow compatibility
    with open(os.path.join(out_dir, "summary.md"), "w") as f:
        f.write(md)
    
    # Write to GITHUB_STEP_SUMMARY if running in Actions
    if "GITHUB_STEP_SUMMARY" in os.environ:
        with open(os.environ["GITHUB_STEP_SUMMARY"], "a") as f:
            f.write(md + "\n")
    
    print(f"\n✅ All reports generated successfully!")
    print(f"   📁 Output: {out_dir}")
    print(f"   📊 Results: {total} total, {passed} passed, {failed} failed")


if __name__ == '__main__':
    main()
